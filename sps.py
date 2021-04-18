import random
import openpyxl
import datetime
from os import path

options = ["scissors", "paper", "stone"]

#Writes score of game in excel sheet
def writeExcelSheet(user_score, computer_score, username):

    #getting date and also the result of the game
    date = (str(datetime.datetime.now()).split("."))[0]
    result = ""
    if(user_score > computer_score):
        result = username + " won"
    elif(user_score == computer_score):
        result = "Draw"
    else:
        result = "nullPointer won"

    #checking if the excel file exists
    if(path.exists("score.xlsx")):
        work_book = openpyxl.load_workbook("score.xlsx")
        work_sheet = work_book.active
        work_sheet.append([date, result, user_score, computer_score])
        work_book.save("score.xlsx")
    else:
        #if no then create one put the headings and recall this function
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(1,1).value = "Date_Time"
        sheet.cell(1,2).value = "Result"
        sheet.cell(1,3).value = "UserScore"
        sheet.cell(1,4).value = "ComputerScore"
        wb.save("score.xlsx")
        writeExcelSheet(user_score, computer_score, username)


def computerChoose():
    choice = random.randint(0, len(options)-1)
    return options[choice]
